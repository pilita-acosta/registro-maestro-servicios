from datetime import date
from pathlib import Path
from uuid import uuid4

from fastapi import Depends, FastAPI, File, HTTPException, Query, UploadFile, status
from openpyxl import load_workbook
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy import or_, select
from sqlalchemy.orm import Session, joinedload

from .database import Base, SessionLocal, engine, get_db, migrate_legacy_schema
from .models import (
    Account,
    AuditEvent,
    Contract,
    CompensationSummary,
    Dependency,
    DependencySecretaryAssignment,
    Domicile,
    ExpectedInvoice,
    InvoiceAllocation,
    InvoiceClaim,
    ImportBatch,
    Invoice,
    Liquidation,
    RentalProperty,
    PhoneLine,
    MobileDevice,
    PhoneIncident,
    Provider,
    PaymentResolution,
    Secretary,
    ServicePoint,
    ServiceType,
)
from .schemas import (
    AccountCreate,
    AccountRead,
    AssignmentCreate,
    ContractCreate,
    ContractRead,
    DependencyCreate,
    DependencyRead,
    DomicileCreate,
    DomicileRead,
    ExpectedInvoiceRead,
    ProviderCreate,
    ProviderRead,
    SecretaryCreate,
    SecretaryRead,
    ServicePointCreate,
    ServicePointRead,
    ServiceTypeCreate,
    ServiceTypeRead,
    InvoiceCreate,
    InvoiceRead,
    InvoiceClaimCreate,
    InvoiceClaimRead,
    InvoiceReceiveCreate,
    InvoiceAllocationRead,
    LiquidationCreate,
    LiquidationRead,
    AuditEventRead,
    PaymentResolutionCreate,
    PaymentResolutionRead,
    CompensationSummaryCreate,
    CompensationSummaryRead,
    RentalPropertyCreate,
    RentalPropertyRead,
    PhoneLineCreate,
    PhoneLineRead,
    MobileDeviceCreate,
    MobileDeviceRead,
    PhoneIncidentCreate,
    PhoneIncidentRead,
    ImportBatchRead,
)

app = FastAPI(
    title="Registro Maestro de Servicios Municipales",
    version="0.1.0",
    description="API para consultar y mantener el padrón maestro de puntos de servicio.",
)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def record_audit(database: Session, entity_type: str, entity_id: int, action: str, details: str) -> None:
    database.add(AuditEvent(entity_type=entity_type, entity_id=entity_id, action=action, details=details))


@app.on_event("startup")
def create_tables() -> None:
    Base.metadata.create_all(bind=engine)
    migrate_legacy_schema()
    seed_demo_data()


def seed_demo_data() -> None:
    database = SessionLocal()
    try:
        if database.scalar(select(ServicePoint.id).limit(1)) is not None:
            return
        secretary = Secretary(name="Secretaría de Hacienda")
        domicile = Domicile(address="Calle 12 1234, La Plata")
        provider = Provider(name="EDELAP")
        service_type = ServiceType(name="Electricidad")
        database.add_all([secretary, domicile, provider, service_type])
        database.flush()
        dependency = Dependency(code="TORRE-1", name="Torre 1 - Oficinas", domicile_id=domicile.id)
        contract = Contract(
            code="CT-DEMO-2026",
            start_date=date(2026, 1, 1),
            billing_frequency="mensual",
            provider_id=provider.id,
            service_type_id=service_type.id,
        )
        database.add_all([dependency, contract])
        database.flush()
        assignment = DependencySecretaryAssignment(
            dependency_id=dependency.id,
            secretary_id=secretary.id,
            starts_at=date(2026, 1, 1),
            changed_by="demo",
        )
        point = ServicePoint(
            code="PS-TORRE-1-ENERGIA",
            name="Torre 1 - Servicio eléctrico",
            dependency_id=dependency.id,
            domicile_id=domicile.id,
        )
        database.add_all([assignment, point])
        database.flush()
        account = Account(
            external_code="EDE-458921",
            nis="NIS-458921",
            service_point_id=point.id,
            provider_id=provider.id,
            service_type_id=service_type.id,
            contract_id=contract.id,
        )
        database.add(account)
        database.flush()
        database.add_all([
            ExpectedInvoice(
                period="2026-06",
                due_date=date(2026, 7, 10),
                delivery_deadline=date(2026, 7, 3),
                status="recibida",
                responsible="Oficina Torre 1",
                cutoff_risk="bajo",
                account_id=account.id,
            ),
            ExpectedInvoice(
                period="2026-07",
                due_date=date(2026, 8, 10),
                delivery_deadline=date(2026, 8, 3),
                status="faltante",
                responsible="Oficina Torre 1",
                cutoff_risk="alto",
                account_id=account.id,
            ),
        ])
        database.commit()
    finally:
        database.close()


@app.get("/health", tags=["sistema"])
def health() -> dict[str, str]:
    return {"status": "ok"}


@app.get("/api/v1/reportes/resumen", tags=["reportes"])
def report_summary(database: Session = Depends(get_db)) -> dict[str, object]:
    expected = list(database.scalars(select(ExpectedInvoice)))
    invoices = list(database.scalars(select(Invoice)))
    return {
        "expected": len(expected),
        "missing": sum(item.status in {"faltante", "reclamada"} for item in expected),
        "received": sum(item.status not in {"faltante", "reclamada"} for item in expected),
        "invoices": len(invoices),
        "invoiced_amount": sum(item.amount for item in invoices),
        "paid_amount": sum(
            resolution.amount
            for invoice in invoices
            for resolution in invoice.payment_resolutions
        ),
        "open_incidents": sum(item.status != "cerrada" for item in database.scalars(select(PhoneIncident))),
    }


@app.post("/api/v1/importaciones/previsualizar", response_model=ImportBatchRead, status_code=status.HTTP_201_CREATED, tags=["importaciones"])
def preview_import(file: UploadFile = File(...), database: Session = Depends(get_db)):
    if not file.filename or Path(file.filename).suffix.lower() not in {".xlsx", ".xlsm"}:
        raise HTTPException(status_code=415, detail="La previsualización requiere un archivo Excel .xlsx o .xlsm")
    safe_name = Path(file.filename).name
    target = Path("/app/uploads") / f"{uuid4().hex}_{safe_name}"
    content = file.file.read()
    if len(content) > 25 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="El archivo supera el límite de 25 MB")
    target.write_bytes(content)
    try:
        workbook = load_workbook(target, read_only=True, data_only=True)
        sheet_count = len(workbook.sheetnames)
        row_count = 0
        valid_row_count = 0
        error_row_count = 0
        warnings: list[str] = []
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [str(value).strip() if value is not None else "" for value in rows[0]]
            data_rows = rows[1:]
            row_count += len(data_rows)
            for row in data_rows:
                if not any(value is not None and str(value).strip() for value in row):
                    continue
                if len(row) >= 2 and row[0] is not None and row[1] is not None:
                    valid_row_count += 1
                else:
                    error_row_count += 1
            if not headers or all(not header for header in headers):
                warnings.append(f"La hoja {sheet.title} no tiene encabezados reconocibles")
        workbook.close()
    except Exception as error:
        target.unlink(missing_ok=True)
        raise HTTPException(status_code=422, detail=f"No se pudo leer el Excel: {error}") from error
    if error_row_count:
        warnings.append(f"{error_row_count} fila(s) requieren revisión antes de confirmar")
    batch = ImportBatch(
        filename=safe_name,
        original_path=str(target),
        sheet_count=sheet_count,
        row_count=row_count,
        valid_row_count=valid_row_count,
        error_row_count=error_row_count,
        warnings="; ".join(warnings) if warnings else None,
    )
    database.add(batch)
    database.commit()
    database.refresh(batch)
    return batch


@app.get("/api/v1/bandeja-facturas", tags=["facturas"])
def invoice_inbox(database: Session = Depends(get_db)) -> list[dict[str, object]]:
    expected = database.execute(
        select(ExpectedInvoice)
        .options(joinedload(ExpectedInvoice.account).joinedload(Account.service_point), joinedload(ExpectedInvoice.account).joinedload(Account.invoices).joinedload(Invoice.payment_resolutions))
        .order_by(ExpectedInvoice.delivery_deadline)
    ).unique().scalars().all()
    return [
        {
            "id": item.id,
            "period": item.period,
            "due_date": item.due_date,
            "delivery_deadline": item.delivery_deadline,
            "status": item.status,
            "responsible": item.responsible,
            "cutoff_risk": item.cutoff_risk,
            "account_id": item.account_id,
            "account": item.account.external_code,
            "point": item.account.service_point.name,
            "invoice_id": next((invoice.id for invoice in item.account.invoices if invoice.period == item.period), None),
            "invoice_status": next((invoice.status for invoice in item.account.invoices if invoice.period == item.period), None),
            "invoice_amount": next((invoice.amount for invoice in item.account.invoices if invoice.period == item.period), 0),
            "resolved_amount": sum(sum(resolution.amount for resolution in invoice.payment_resolutions) for invoice in item.account.invoices if invoice.period == item.period),
        }
        for item in expected
    ]


@app.post("/api/v1/facturas", response_model=InvoiceRead, status_code=status.HTTP_201_CREATED, tags=["facturas"])
def create_invoice(payload: InvoiceCreate, database: Session = Depends(get_db)):
    item = Invoice(**payload.model_dump())
    database.add(item)
    database.commit()
    database.refresh(item)
    return item


@app.post("/api/v1/facturas/{item_id}/liquidar", response_model=LiquidationRead, status_code=status.HTTP_201_CREATED, tags=["liquidaciones"])
def liquidate_invoice(item_id: int, payload: LiquidationCreate, database: Session = Depends(get_db)):
    invoice = database.get(Invoice, item_id)
    secretary = database.get(Secretary, payload.secretary_id)
    if invoice is None:
        raise HTTPException(status_code=404, detail="Factura no encontrada")
    if secretary is None:
        raise HTTPException(status_code=404, detail="Secretaría no encontrada")
    if invoice.liquidation_id is not None:
        raise HTTPException(status_code=409, detail="La factura ya está liquidada")
    liquidation = Liquidation(
        code=f"LIQ-{invoice.period}-{invoice.id}",
        period=invoice.period,
        resolution_mode=payload.resolution_mode,
    )
    database.add(liquidation)
    database.flush()
    invoice.liquidation_id = liquidation.id
    invoice.status = "aprobada_para_liquidar"
    database.add(InvoiceAllocation(invoice_id=invoice.id, secretary_id=secretary.id, amount=invoice.amount))
    record_audit(database, "factura", invoice.id, "liquidada", f"Liquidación {liquidation.code}; Secretaría {secretary.name}")
    database.commit()
    database.refresh(liquidation)
    return liquidation


@app.post("/api/v1/facturas-esperadas/{item_id}/recibir", response_model=InvoiceRead, status_code=status.HTTP_201_CREATED, tags=["facturas"])
def receive_expected_invoice(item_id: int, payload: InvoiceReceiveCreate, database: Session = Depends(get_db)):
    item = database.get(ExpectedInvoice, item_id)
    if item is None:
        raise HTTPException(status_code=404, detail="Factura esperada no encontrada")
    duplicate = database.scalar(
        select(Invoice).where(Invoice.number == payload.number, Invoice.account_id == item.account_id)
    )
    if duplicate is not None:
        raise HTTPException(status_code=409, detail="La factura ya está registrada para esta cuenta")
    invoice = Invoice(
        number=payload.number,
        period=item.period,
        amount=payload.amount,
        issued_at=payload.issued_at,
        due_date=payload.due_date or item.due_date,
        resolution_mode=payload.resolution_mode,
        original_file=payload.original_file,
        account_id=item.account_id,
    )
    item.status = "recibida"
    item.cutoff_risk = "bajo"
    database.add(invoice)
    database.flush()
    record_audit(database, "factura", invoice.id, "recibida", f"Número {invoice.number}; importe {invoice.amount}")
    database.commit()
    database.refresh(invoice)
    return invoice


@app.post("/api/v1/facturas-esperadas/{item_id}/reclamos", response_model=InvoiceClaimRead, status_code=status.HTTP_201_CREATED, tags=["facturas"])
def create_invoice_claim(item_id: int, payload: InvoiceClaimCreate, database: Session = Depends(get_db)):
    expected = database.get(ExpectedInvoice, item_id)
    if expected is None:
        raise HTTPException(status_code=404, detail="Factura esperada no encontrada")
    claim = InvoiceClaim(expected_invoice_id=item_id, **payload.model_dump())
    expected.status = "reclamada"
    database.add(claim)
    record_audit(database, "factura_esperada", expected.id, "reclamo_creado", payload.message)
    database.commit()
    database.refresh(claim)
    return claim


@app.get("/api/v1/facturas/{item_id}/auditoria", response_model=list[AuditEventRead], tags=["auditoria"])
def invoice_audit(item_id: int, database: Session = Depends(get_db)):
    if database.get(Invoice, item_id) is None:
        raise HTTPException(status_code=404, detail="Factura no encontrada")
    return list(database.scalars(
        select(AuditEvent)
        .where(AuditEvent.entity_type == "factura", AuditEvent.entity_id == item_id)
        .order_by(AuditEvent.created_at)
    ))


@app.post("/api/v1/facturas/{item_id}/resolver", response_model=PaymentResolutionRead, status_code=status.HTTP_201_CREATED, tags=["pagos"])
def resolve_invoice(item_id: int, payload: PaymentResolutionCreate, database: Session = Depends(get_db)):
    invoice = database.get(Invoice, item_id)
    if invoice is None:
        raise HTTPException(status_code=404, detail="Factura no encontrada")
    if invoice.liquidation_id is None:
        raise HTTPException(status_code=409, detail="La factura debe estar liquidada antes de resolver el pago")
    resolved_amount = sum(resolution.amount for resolution in invoice.payment_resolutions)
    if resolved_amount + payload.amount > invoice.amount:
        raise HTTPException(status_code=422, detail="El importe resuelto supera el importe de la factura")
    resolution = PaymentResolution(invoice_id=invoice.id, **payload.model_dump())
    database.add(resolution)
    database.flush()
    total = resolved_amount + payload.amount
    invoice.status = "compensada" if payload.mode == "compensacion_proveedor" and total == invoice.amount else ("pagada" if total == invoice.amount else "pago_parcial")
    record_audit(database, "factura", invoice.id, "resolucion_registrada", f"{payload.mode}; importe {payload.amount}; destinatario {payload.recipient}")
    database.commit()
    database.refresh(resolution)
    return resolution


@app.post("/api/v1/resumenes-compensacion/{summary_id}/facturas/{invoice_id}", response_model=CompensationSummaryRead, tags=["compensaciones"])
def attach_invoice_to_compensation(summary_id: int, invoice_id: int, database: Session = Depends(get_db)):
    summary = database.get(CompensationSummary, summary_id)
    invoice = database.get(Invoice, invoice_id)
    if summary is None or invoice is None:
        raise HTTPException(status_code=404, detail="Resumen o factura no encontrada")
    if invoice.account.provider_id != summary.provider_id:
        raise HTTPException(status_code=422, detail="La factura no pertenece al proveedor del resumen")
    if invoice.period != summary.period:
        raise HTTPException(status_code=422, detail="La factura no pertenece al período del resumen")
    if invoice.status not in {"aprobada_para_liquidar", "pago_parcial", "compensada", "pagada"}:
        raise HTTPException(status_code=409, detail="La factura todavía no está controlada para compensar")
    if invoice.compensation_summary_id not in {None, summary.id}:
        raise HTTPException(status_code=409, detail="La factura ya está asociada a otro resumen")
    invoice.compensation_summary_id = summary.id
    summary.validated_amount = sum(item.amount for item in summary.invoices) + invoice.amount
    summary.status = "observado" if summary.validated_amount > summary.reported_amount else "validado"
    record_audit(database, "resumen_compensacion", summary.id, "factura_asociada", f"Factura {invoice.number}")
    database.commit()
    database.refresh(summary)
    return summary


@app.get("/api/v1/puntos-servicio", tags=["puntos de servicio"])
def search_service_points(
    query: str | None = Query(default=None, min_length=1),
    database: Session = Depends(get_db),
) -> dict[str, object]:
    statement = (
        select(ServicePoint)
        .outerjoin(ServicePoint.domicile)
        .outerjoin(ServicePoint.dependency)
        .outerjoin(ServicePoint.accounts)
        .outerjoin(Account.provider)
    ).order_by(ServicePoint.name).distinct()
    if query:
        pattern = f"%{query.strip()}%"
        statement = statement.where(
            or_(
                ServicePoint.code.ilike(pattern),
                ServicePoint.name.ilike(pattern),
                Domicile.address.ilike(pattern),
                Dependency.name.ilike(pattern),
                Account.external_code.ilike(pattern),
                Account.nis.ilike(pattern),
                Provider.name.ilike(pattern),
            )
        )
    items = list(database.scalars(statement))
    response_items = [
        ServicePointRead.model_validate(item).model_copy(update={"address": item.domicile.address})
        for item in items
    ]
    return {"items": response_items, "query": query, "total": len(items)}


@app.get("/api/v1/puntos-servicio/{item_id}/detalle", tags=["puntos de servicio"])
def service_point_detail(item_id: int, database: Session = Depends(get_db)):
    point = database.execute(
        select(ServicePoint)
        .options(
            joinedload(ServicePoint.domicile),
            joinedload(ServicePoint.dependency).joinedload(Dependency.assignments).joinedload(DependencySecretaryAssignment.secretary),
            joinedload(ServicePoint.accounts).joinedload(Account.provider),
            joinedload(ServicePoint.accounts).joinedload(Account.service_type),
            joinedload(ServicePoint.accounts).joinedload(Account.contract),
        )
        .where(ServicePoint.id == item_id)
    ).unique().scalar_one_or_none()
    if point is None:
        raise HTTPException(status_code=404, detail="Punto de servicio no encontrado")
    return {
        "id": point.id,
        "code": point.code,
        "name": point.name,
        "status": point.status,
        "address": point.domicile.address,
        "dependency": point.dependency.name if point.dependency else None,
        "secretaries": [assignment.secretary.name for assignment in (point.dependency.assignments if point.dependency else [])],
        "accounts": [
            {
                "id": account.id,
                "external_code": account.external_code,
                "nis": account.nis,
                "provider": account.provider.name,
                "service": account.service_type.name,
                "contract": account.contract.code,
            }
            for account in point.accounts
        ],
    }


def create_crud_routes(create_schema, read_schema, model, route: str):
    @app.post(route, response_model=read_schema, status_code=status.HTTP_201_CREATED, tags=[route.strip("/")])
    def create_item(payload: create_schema, database: Session = Depends(get_db)):
        item = model(**payload.model_dump())
        database.add(item)
        database.commit()
        database.refresh(item)
        return item

    @app.get(route, response_model=list[read_schema], tags=[route.strip("/")])
    def list_items(database: Session = Depends(get_db)):
        return list(database.scalars(select(model).order_by(model.id)))


create_crud_routes(SecretaryCreate, SecretaryRead, Secretary, "/api/v1/secretarias")
create_crud_routes(DomicileCreate, DomicileRead, Domicile, "/api/v1/domicilios")
create_crud_routes(RentalPropertyCreate, RentalPropertyRead, RentalProperty, "/api/v1/inmuebles-alquilados")
create_crud_routes(DependencyCreate, DependencyRead, Dependency, "/api/v1/dependencias")
create_crud_routes(ProviderCreate, ProviderRead, Provider, "/api/v1/proveedores")
create_crud_routes(ServiceTypeCreate, ServiceTypeRead, ServiceType, "/api/v1/servicios")
create_crud_routes(ContractCreate, ContractRead, Contract, "/api/v1/contratos")
create_crud_routes(CompensationSummaryCreate, CompensationSummaryRead, CompensationSummary, "/api/v1/resumenes-compensacion")
create_crud_routes(PhoneLineCreate, PhoneLineRead, PhoneLine, "/api/v1/lineas-telefonicas")
create_crud_routes(MobileDeviceCreate, MobileDeviceRead, MobileDevice, "/api/v1/dispositivos-moviles")
create_crud_routes(PhoneIncidentCreate, PhoneIncidentRead, PhoneIncident, "/api/v1/incidencias-telefonia")


@app.patch("/api/v1/incidencias-telefonia/{item_id}/cerrar", response_model=PhoneIncidentRead, tags=["telefonía"])
def close_phone_incident(item_id: int, resolution: str = Query(min_length=3), database: Session = Depends(get_db)):
    incident = database.get(PhoneIncident, item_id)
    if incident is None:
        raise HTTPException(status_code=404, detail="Incidencia no encontrada")
    incident.status = "cerrada"
    incident.resolution = resolution
    incident.closed_at = date.today()
    database.commit()
    database.refresh(incident)
    return incident


@app.post("/api/v1/puntos-servicio", response_model=ServicePointRead, status_code=status.HTTP_201_CREATED, tags=["puntos de servicio"])
def create_service_point(payload: ServicePointCreate, database: Session = Depends(get_db)):
    item = ServicePoint(**payload.model_dump())
    database.add(item)
    database.commit()
    database.refresh(item)
    return item


@app.get("/api/v1/puntos-servicio/{item_id}", response_model=ServicePointRead, tags=["puntos de servicio"])
def get_service_point(item_id: int, database: Session = Depends(get_db)):
    item = database.get(ServicePoint, item_id)
    if item is None:
        raise HTTPException(status_code=404, detail="Punto de servicio no encontrado")
    return item


@app.post("/api/v1/cuentas", response_model=AccountRead, status_code=status.HTTP_201_CREATED, tags=["cuentas"])
def create_account(payload: AccountCreate, database: Session = Depends(get_db)):
    item = Account(**payload.model_dump())
    database.add(item)
    database.commit()
    database.refresh(item)
    return item


@app.post("/api/v1/asignaciones-dependencia-secretaria", status_code=status.HTTP_201_CREATED, tags=["asignaciones"])
def create_assignment(payload: AssignmentCreate, database: Session = Depends(get_db)):
    item = DependencySecretaryAssignment(**payload.model_dump())
    database.add(item)
    database.commit()
    database.refresh(item)
    return item
